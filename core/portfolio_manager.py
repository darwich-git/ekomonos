import os
import glob

class PortfolioManager:
    def __init__(self, root_dir):
        self.root_dir = root_dir
        self.companies = self._scan_companies()

    def _scan_companies(self):
        """Scans the root directory for company folders."""
        companies = {}
        if not os.path.exists(self.root_dir):
            return companies

        # Look for folders in the STOCK directory (tickers are folder names)
        for item in os.listdir(self.root_dir):
            item_path = os.path.join(self.root_dir, item)
            if os.path.isdir(item_path):
                ticker = item # New structure: Folder name IS the ticker
                excel_file = self._find_excel_file(item_path, ticker)
                completeness = self._calculate_completeness(item_path, ticker, excel_file)
                
                companies[ticker] = {
                    "path": item_path,
                    "excel_file": excel_file,
                    "completeness": completeness
                }
        return companies

    def _calculate_completeness(self, folder_path, ticker, excel_file):
        """
        Calculates data completeness % based on:
        - 5 Annual Reports (approx 50%)
        - 1 Excel Model (30%)
        - 1 Pre-filter/Checklist (20%)
        """
        score = 0
        
        # 1. Excel Model
        if excel_file:
            score += 30
            
        # 2. Checklist (look for 'Checklist' in filename in VARIOS or root)
        has_checklist = False
        varios_path = os.path.join(folder_path, f"4 VARIOS {ticker}")
        
        # Check root and Varios
        paths_to_check = [folder_path]
        if os.path.exists(varios_path):
            paths_to_check.append(varios_path)
            # Also check year subfolders in Varios
            for item in os.listdir(varios_path):
                sub = os.path.join(varios_path, item)
                if os.path.isdir(sub):
                    paths_to_check.append(sub)

        for p in paths_to_check:
            if not os.path.exists(p): continue
            for f in os.listdir(p):
                if "checklist" in f.lower():
                    has_checklist = True
                    break
            if has_checklist: break
            
        if has_checklist:
            score += 20
            
        # 3. Annual Reports (look in 1 REPORTS <TICKER>)
        report_count = 0
        reports_path = os.path.join(folder_path, f"1 REPORTS {ticker}")
        
        if os.path.exists(reports_path):
            for root, dirs, files in os.walk(reports_path):
                for f in files:
                    if f.lower().endswith(".pdf"):
                        report_count += 1
        
        # Cap at 5 reports
        reports_score = min(report_count, 5) * 10
        score += reports_score
        
        return score

    def _find_excel_file(self, folder_path, ticker):
        """Finds the main Excel model file in the folder."""
        # Look in 3 EXCEL <TICKER>
        excel_dir = os.path.join(folder_path, f"3 EXCEL {ticker}")
        
        candidates = []
        if os.path.exists(excel_dir):
            candidates = glob.glob(os.path.join(excel_dir, "*.xls*"))
        
        # If empty, maybe in root?
        if not candidates:
            candidates = glob.glob(os.path.join(folder_path, "*.xls*"))
            
        if not candidates:
            return None
            
        # Priority 1: Contains "Modelo"
        for f in candidates:
            if "Modelo" in os.path.basename(f):
                return f
                
        # Priority 2: Most recently modified file
        candidates.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        return candidates[0]

    def get_company_data(self, ticker, sheet_name="1.Income statement"):
        """Reads data from the specified company Excel file."""
        if ticker not in self.companies or not self.companies[ticker]["excel_file"]:
            return None

        file_path = self.companies[ticker]["excel_file"]
        
        try:
            # Read with header=None to handle the custom layout
            # We know data starts around row 7 (0-indexed) based on inspection
            import pandas as pd
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            return df
        except Exception as e:
            print(f"Error reading {ticker}: {e}")
            return None

    def update_cell(self, ticker, sheet_name, row, col, value):
        """Updates a specific cell in the Excel file."""
        if ticker not in self.companies or not self.companies[ticker]["excel_file"]:
            return False

        file_path = self.companies[ticker]["excel_file"]
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, keep_vba=True)
            if sheet_name not in wb.sheetnames:
                return False
                
            ws = wb[sheet_name]
            # openpyxl is 1-indexed, pandas is 0-indexed
            # If row/col come from pandas dataframe index, add 1
            ws.cell(row=row+1, column=col+1, value=value)
            
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"Error updating {ticker}: {e}")
            return False

    def get_sheet_names(self, ticker):
        if ticker not in self.companies or not self.companies[ticker]["excel_file"]:
            return []
        
        file_path = self.companies[ticker]["excel_file"]
        try:
            import pandas as pd
            xls = pd.ExcelFile(file_path)
            return xls.sheet_names
        except:
            return []
