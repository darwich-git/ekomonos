import openpyxl
import pandas as pd
import datetime
from core.database import MonthlySnapshot, AccountBalance, IncomeRecord, Account, DB_PATH
from sqlalchemy import create_engine
from config import MASTER_BALANCE_PATH

def map_month_id_to_es_str(month_id):
    # '2026-01' -> 'Enero/26'
    y, m = month_id.split('-')
    months = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
    }
    return f"{months[m]}/{y[-2:]}"

def normalize_str(s):
    import unicodedata
    if not s: return ""
    return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode('utf-8').lower()

def safe_float(val):
    if val is None: return 0.0
    try: return float(val)
    except: return 0.0

def export_snapshot_to_master(snapshot_id):
    from core.database import get_session
    session = get_session()
    
    snap = session.query(MonthlySnapshot).get(snapshot_id)
    if not snap:
        session.close()
        return False
        
    bals = session.query(AccountBalance).filter_by(snapshot_id=snap.id).all()
    inc_rafa = session.query(IncomeRecord).filter_by(month_id=snap.month_id, owner="Rafa").first()
    inc_cris = session.query(IncomeRecord).filter_by(month_id=snap.month_id, owner="Cris").first()
    
    # 1. Load Excel
    file_path = str(MASTER_BALANCE_PATH)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error opening Excel: {e}")
        session.close()
        return False
        
    sheet = wb["balance"] if "balance" in wb.sheetnames else wb.active
    
    # 2. Find or Create Column
    tar_month = map_month_id_to_es_str(snap.month_id)
    col_idx = -1
    for c in range(1, 100):
        val = sheet.cell(row=1, column=c).value
        # Check matching string or datetime
        if isinstance(val, str) and normalize_str(val) == normalize_str(tar_month):
            col_idx = c
            break
        elif isinstance(val, datetime.datetime):
            # 'Mayo/20' is sometimes parsed as datetime
            # We'll just append if we don't find a direct text match, or look for last col
            pass
            
    if col_idx == -1:
        # Find first empty column in row 1
        for c in range(2, 100):
            if sheet.cell(row=1, column=c).value is None:
                col_idx = c
                break
        if col_idx == -1: col_idx = 100 # Fallback
        
        sheet.cell(row=1, column=col_idx).value = tar_month

    # Prepare values mapping
    val_map = {}
    
    for b in bals:
        acc = session.query(Account).get(b.account_id)
        if not acc: continue
        
        norm_name = normalize_str(acc.name)
        v_cur = safe_float(b.current_value)
        v_ap = safe_float(b.invested_amount)
        v_pr = v_cur - v_ap
        
        if 'ibkr' in norm_name and 'rafa' in norm_name:
            val_map['aportado ibkr rafa'] = v_ap
            val_map['profit ibkr rafa'] = v_pr
        elif 'fundsmith' in norm_name and 'rafa' in norm_name:
            val_map['aportado fundsmith rafa'] = v_ap
            val_map['profit fundsmith rafa'] = v_pr
        elif 'fundsmith' in norm_name and 'cris' in norm_name:
            val_map['aportado fundsmith cris'] = v_ap
            val_map['profit fundsmith cris'] = v_pr
        elif 'ibkr' in norm_name and 'cris' in norm_name:
            val_map['aportado ibkr cris'] = v_ap # Might not exist in excel
            val_map['profit ibkr cris'] = v_pr
        elif 'monetario' in norm_name and 'cris' in norm_name:
            val_map['fondo monetario cris'] = v_cur # No profit split in excel
        else:
            # Bank accounts or debts
            if 'formacion' in norm_name: val_map['cuenta formacion rafa'] = v_cur
            elif 'inversion' in norm_name and 'rafa' in norm_name: val_map['cuenta inversion rafa'] = v_cur
            elif 'insta' in norm_name: val_map['cuenta ahorro (instan saving) cris'] = v_cur
            elif 'monthly' in norm_name: val_map['cuenta ahorro (monthly saving) cris'] = v_cur
            elif 'caixa' in norm_name: val_map['cuenta la caixa cris (eur)'] = v_cur
            elif 'bankia' in norm_name: val_map['cuenta bankia rafa (eur)'] = v_cur
            elif 'hamburgh' in norm_name and acc.type == 'real_estate': val_map['casa hamburgh place'] = v_cur
            elif 'hamburgh' in norm_name and acc.type == 'mortgage': val_map['hipoteca hamburgh place'] = v_cur
            elif 'deuda equity' in norm_name or 'interna rafa' in norm_name: val_map['deuda interna rafa'] = v_cur
            else:
                val_map[norm_name] = v_cur

    # Add Income/Exp
    if inc_rafa:
        val_map['ingresos brutos rafa'] = safe_float(inc_rafa.gross_amount)
        val_map['ingresos netos rafa'] = safe_float(inc_rafa.net_amount)
        val_map['gastos individuales rafa'] = safe_float(inc_rafa.total_expenses)
        val_map['cuota rafa a familia'] = safe_float(inc_rafa.shared_quota)
    if inc_cris:
        val_map['ingresos brutos cris'] = safe_float(inc_cris.gross_amount)
        val_map['ingresos netos cris'] = safe_float(inc_cris.net_amount)
        val_map['gastos individuales cris'] = safe_float(inc_cris.total_expenses)
        val_map['cuota cris a familia'] = safe_float(inc_cris.shared_quota)
        
    val_map['gastos familia (comun)'] = safe_float(snap.comun_expenses)
    
    # Optional: Notas
    if snap.notes:
        val_map['notas del mes'] = snap.notes
        
    # Write to Excel
    for row in range(1, 150):
        cell_val = sheet.cell(row=row, column=1).value
        if not cell_val: continue
        
        nc = normalize_str(str(cell_val))
        
        for k, val in val_map.items():
            # Exact or partial match logic
            if k == nc or (k in nc and len(k) > 5) or (nc in k and len(nc) > 5):
                cell = sheet.cell(row=row, column=col_idx)
                cell.value = val
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
                break
                
    try:
        wb.save(file_path)
        session.close()
        return True
    except Exception as e:
        print(f"Failed to save Excel (Is it open?): {e}")
        session.close()
        return False
