import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

def rebuild_excel():
    old_path = r'd:\Proyectos\EKKOMONOS\Master_Balance.xlsx'
    new_path = r'd:\Proyectos\EKKOMONOS\Master_Balance_V2.xlsx'

    # Load old workbook
    old_wb = openpyxl.load_workbook(old_path, data_only=True)
    old_ws = old_wb.active

    # Create new workbook
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = 'balance'

    # Copy the first 3 rows (headers) from old to new exactly
    for row in range(1, 4):
        for col in range(1, old_ws.max_column + 1):
            new_ws.cell(row=row, column=col).value = old_ws.cell(row=row, column=col).value
            # Copy basic styling if needed, but value is most important

    # Define the new structure: (New_Name, Old_Name_To_Search, Is_Header/Style)
    # Old_Name_To_Search is a list of possible matches in the old file
    structure = [
        ("PARÁMETROS MACRO", [], "HEADER_MAIN"),
        ("Cambio GBP a EUR", ["cambio", "cambio gbp"], "NORMAL"),
        ("Cambio USD a EUR", [], "NORMAL"),
        ("", [], "EMPTY"),
        
        ("ACTIVO Y LIQUIDEZ: RAFA", [], "HEADER_MAIN"),
        ("Cuenta Master Rafa", ["Bank of scot. Master"], "NORMAL"),
        ("Cuenta Inversión Rafa", ["Bank of scot. inversion"], "NORMAL"),
        ("Cuenta Formación Rafa", ["Bank of scot. formaci"], "NORMAL"),
        ("Cuenta Bankia Rafa (EUR)", ["Cuenta bankia"], "NORMAL"),
        ("Efectivo Rafa", ["Efectivo"], "NORMAL"),
        ("Aportado IBKR Rafa", ["Aportado IBKR"], "NORMAL"),
        ("Profit IBKR Rafa", ["Profit IBKR"], "NORMAL"),
        ("Total IBKR Rafa", ["Total IBKR"], "FORMULA"),
        ("Aportado Fundsmith Rafa", ["Fondos Fundsmith"], "NORMAL"), # You mentioned "Fondos Fundsmith" had the invested amount, or "Fundsmith aportado"
        ("Profit Fundsmith Rafa", ["Profit Fundsmith"], "NORMAL"),
        ("Total Fundsmith Rafa", ["Fundsmith TOTAL"], "FORMULA"),
        ("TOTAL ACTIVOS RAFA", ["total activos corrientes"], "HEADER_SUB"),
        ("", [], "EMPTY"),
        
        ("ACTIVO Y LIQUIDEZ: CRIS", [], "HEADER_MAIN"),
        ("Cuenta Master Cris", ["Cuenta marter a"], "NORMAL"),
        ("Cuenta Ahorro (Instan saving) Cris", ["instan saving"], "NORMAL"),
        ("Cuenta Ahorro (Monthly saving) Cris", ["monthly saving"], "NORMAL"),
        ("Cuenta La Caixa Cris (EUR)", ["LA caixa a"], "NORMAL"),
        ("Fondo Monetario Cris", ["Plazo fijo chuis"], "NORMAL"),
        ("Aportado IBKR Cris", [], "NORMAL"),
        ("Profit IBKR Cris", ["IBRK CHUIS"], "NORMAL"),
        ("Total IBKR Cris", [], "FORMULA"),
        ("Aportado Fundsmith Cris", ["FundSmith Aportado a"], "NORMAL"),
        ("Profit Fundsmith Cris", ["FundSmith Ganacia a"], "NORMAL"),
        ("Total Fundsmith Cris", [], "FORMULA"),
        ("TOTAL ACTIVOS CRIS", [], "HEADER_SUB"),
        ("", [], "EMPTY"),
        
        ("ACTIVO FAMILIAR E INMOBILIARIO", [], "HEADER_MAIN"),
        ("Cuenta Común", ["Cuenta Comun"], "NORMAL"),
        ("Casa Hamburgh Place", ["2/6 hamburgh place"], "NORMAL"),
        ("Deuda Interna Rafa", [], "NORMAL"),
        ("TOTAL ACTIVOS FAMILIARES", [], "HEADER_SUB"),
        ("", [], "EMPTY"),
        
        ("PASIVOS Y DEUDAS", [], "HEADER_MAIN"),
        ("Hipoteca Cris", ["Hipoteca Cris"], "NORMAL"),
        ("Hipoteca Rafa", ["hipoteca Rafa"], "NORMAL"),
        ("Crédito", ["credito"], "NORMAL"),
        ("TOTAL PASIVOS", [], "HEADER_SUB"),
        ("", [], "EMPTY"),
        
        ("PATRIMONIO NETO", [], "HEADER_MAIN"),
        ("PATRIMONIO NETO RAFA", [], "FORMULA"),
        ("PATRIMONIO NETO CRIS", [], "FORMULA"),
        ("PATRIMONIO NETO FAMILIA", [], "FORMULA"),
        ("PATRIMONIO NETO GLOBAL (LIBRAS)", ["PASIVO NETO libras"], "HEADER_MAIN"),
        ("PATRIMONIO NETO GLOBAL (EUROS)", ["PASIVO NETO euros"], "HEADER_MAIN"),
        ("", [], "EMPTY"),
        
        ("RADIOGRAFÍA DEL CASH FLOW", [], "HEADER_MAIN"),
        ("Ingresos Brutos Rafa", [], "NORMAL"),
        ("Ingresos Netos Rafa", ["Ingresos netos"], "NORMAL"),
        ("Gastos Individuales Rafa", ["Gastos"], "NORMAL"),
        ("Cuota Rafa a Familia", [], "NORMAL"),
        ("Ahorro Libre Rafa", ["Ahorro"], "FORMULA"),
        ("Ingresos Brutos Cris", [], "NORMAL"),
        ("Ingresos Netos Cris", [], "NORMAL"),
        ("Gastos Individuales Cris", [], "NORMAL"),
        ("Cuota Cris a Familia", [], "NORMAL"),
        ("Ahorro Libre Cris", [], "FORMULA"),
        ("Ingresos Cuenta Común", [], "FORMULA"),
        ("Gastos Familia (Común)", ["Gastos comun"], "NORMAL"),
        ("Balance Cuenta Común", [], "FORMULA"),
        ("", [], "EMPTY"),
        
        ("MÉTRICAS Y PERFORMANCE", [], "HEADER_MAIN"),
        ("Tasa de Ahorro Rafa %", [], "FORMULA"),
        ("Tasa de Ahorro Hogar %", [], "FORMULA"),
        ("Rentabilidad Variable Rafa %", [], "FORMULA"),
        ("Liquidez (Cash) Rafa %", ["%cash"], "FORMULA"),
        ("Crecimiento vs Mes Anterior", ["Credimiento"], "FORMULA"),
        ("", [], "EMPTY"),
        
        ("BITÁCORA", [], "HEADER_MAIN"),
        ("Notas del Mes", [], "NORMAL"),
    ]

    # Map old rows
    old_row_map = {} # old header string -> row index in old worksheet
    for r in range(1, old_ws.max_row + 1):
        val = old_ws.cell(row=r, column=1).value
        if val:
            old_row_map[str(val).lower().strip()] = r
            
    def find_old_row(search_terms):
        for term in search_terms:
            t_lower = term.lower()
            for old_name, r_idx in old_row_map.items():
                if t_lower in old_name:
                    return r_idx
        return None

    # Write new structure starting from row 4
    current_row = 4
    
    # Styling
    header_main_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_sub_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    font_bold_white = Font(bold=True, color="FFFFFF")
    font_bold = Font(bold=True)

    for new_name, search_terms, row_type in structure:
        new_ws.cell(row=current_row, column=1).value = new_name
        
        # Apply styles
        cell_A = new_ws.cell(row=current_row, column=1)
        if row_type == "HEADER_MAIN":
            cell_A.fill = header_main_fill
            cell_A.font = font_bold_white
        elif row_type == "HEADER_SUB" or row_type == "FORMULA":
            cell_A.fill = header_sub_fill
            cell_A.font = font_bold
            
        # Copy data if not empty or header
        if row_type not in ["HEADER_MAIN", "EMPTY"] and search_terms:
            old_r = find_old_row(search_terms)
            if old_r:
                for col in range(2, old_ws.max_column + 1):
                    new_ws.cell(row=current_row, column=col).value = old_ws.cell(row=old_r, column=col).value

        current_row += 1

    # Format column A width
    new_ws.column_dimensions['A'].width = 40
    
    # Fix first 3 rows
    new_ws.cell(row=1, column=1).value = "AÑO"
    new_ws.cell(row=2, column=1).value = "DIVISOR"
    new_ws.cell(row=3, column=1).value = "CONCEPTO"
    new_ws.cell(row=1, column=1).font = font_bold
    new_ws.cell(row=2, column=1).font = font_bold
    new_ws.cell(row=3, column=1).font = font_bold
    
    # Save
    new_wb.save(new_path)
    print(f"Generated successfully: {new_path}")

if __name__ == '__main__':
    rebuild_excel()
