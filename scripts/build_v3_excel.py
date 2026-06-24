import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re

def parse_old_data():
    old_path = r'd:\Proyectos\EKKOMONOS\Master_Balance.xlsx'
    old_wb = openpyxl.load_workbook(old_path, data_only=True)
    old_ws = old_wb.active

    # Parse headers from Row 1
    months_map = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
        'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'octrubre': 10, 'noviembre': 11, 'noviembrte': 11, 'diciembre': 12
    }
    
    # Identify month columns
    cols_data = [] # (old_col_idx, year, month, old_header_str)
    for col in range(2, old_ws.max_column + 1):
        raw_val = old_ws.cell(row=1, column=col).value
        if not raw_val: continue
        val = str(raw_val).strip().lower()
        
        if '/' in val:
            parts = val.split('/')
            if len(parts) == 2:
                m_str, y_str = parts[0].strip(), parts[1].strip()
                if m_str in months_map and y_str.isdigit():
                    m_num = months_map[m_str]
                    y_num = 2000 + int(y_str) if len(y_str) <= 2 else int(y_str)
                    
                    # Original name formatted nicely
                    nice_name = f"{m_str.capitalize()}/{y_str}"
                    cols_data.append((col, y_num, m_num, nice_name))

    # Sort chronological
    cols_data.sort(key=lambda x: (x[1], x[2]))
    
    # Map old rows
    old_row_map = {}
    for r in range(1, old_ws.max_row + 1):
        val = old_ws.cell(row=r, column=1).value
        if val:
            old_row_map[str(val).lower().strip()] = r
            
    return old_ws, cols_data, old_row_map

def find_old_row(search_terms, old_row_map):
    for term in search_terms:
        t_lower = term.lower()
        for old_name, r_idx in old_row_map.items():
            if t_lower in old_name:
                return r_idx
    return None

def build_v3():
    old_ws, cols_data, old_row_map = parse_old_data()
    new_path = r'd:\Proyectos\EKKOMONOS\Master_Balance_V3.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'balance'

    # The New Structure - including Type of data to format
    # "CURRENCY", "PERCENT", "NORMAL", "TEXT"
    structure = [
        ("🛡️ ACTIVOS BANCARIOS", [], "HEADER_MAIN", "TEXT"),
        ("Cuenta Master Rafa", ["Bank of scot. Master"], "NORMAL", "CURRENCY"),
        ("Cuenta Inversión Rafa", ["Bank of scot. inversion"], "NORMAL", "CURRENCY"),
        ("Cuenta Formación Rafa", ["Bank of scot. formaci"], "NORMAL", "CURRENCY"),
        ("Cuenta Bankia Rafa (EUR)", ["Cuenta bankia"], "NORMAL", "CURRENCY"),
        ("Efectivo Rafa", ["Efectivo"], "NORMAL", "CURRENCY"),
        ("Cuenta Master Cris", ["Cuenta marter a"], "NORMAL", "CURRENCY"),
        ("Cuenta Ahorro (Instan saving) Cris", ["instan saving"], "NORMAL", "CURRENCY"),
        ("Cuenta Ahorro (Monthly saving) Cris", ["monthly saving"], "NORMAL", "CURRENCY"),
        ("Cuenta La Caixa Cris (EUR)", ["LA caixa a"], "NORMAL", "CURRENCY"),
        ("Cuenta Común", ["Cuenta Comun"], "NORMAL", "CURRENCY"),
        ("", [], "EMPTY", ""),
        
        ("📈 INVERSIONES (Rafa)", [], "HEADER_MAIN", "TEXT"),
        ("Aportado IBKR Rafa", ["Aportado IBKR"], "NORMAL", "CURRENCY"),
        ("Profit IBKR Rafa", ["Profit IBKR"], "NORMAL", "CURRENCY"),
        ("Aportado Fundsmith Rafa", ["Fondos Fundsmith"], "NORMAL", "CURRENCY"), 
        ("Profit Fundsmith Rafa", ["Profit Fundsmith"], "NORMAL", "CURRENCY"),
        ("", [], "EMPTY", ""),

        ("📈 INVERSIONES (Cris)", [], "HEADER_MAIN", "TEXT"),
        ("Fondo Monetario Cris", ["Plazo fijo chuis"], "NORMAL", "CURRENCY"),
        ("Aportado IBKR Cris", [], "NORMAL", "CURRENCY"),
        ("Profit IBKR Cris", ["IBRK CHUIS"], "NORMAL", "CURRENCY"),
        ("Aportado Fundsmith Cris", ["FundSmith Aportado a"], "NORMAL", "CURRENCY"),
        ("Profit Fundsmith Cris", ["FundSmith Ganacia a"], "NORMAL", "CURRENCY"),
        ("", [], "EMPTY", ""),
        
        ("📉 PASIVOS E INMOBILIARIO", [], "HEADER_MAIN", "TEXT"),
        ("Casa Hamburgh Place", ["2/6 hamburgh place"], "NORMAL", "CURRENCY"),
        ("Deuda Interna Rafa", [], "NORMAL", "CURRENCY"),
        ("Hipoteca Hamburgh Place", ["hipoteca Rafa"], "NORMAL", "CURRENCY"), # Grouped
        ("Otras Deudas / Créditos", ["credito", "Hipoteca Cris"], "NORMAL", "CURRENCY"),
        ("", [], "EMPTY", ""),
        
        ("🚰 INGRESOS Y GASTOS", [], "HEADER_MAIN", "TEXT"),
        ("Ingresos Brutos Rafa", [], "NORMAL", "CURRENCY"),
        ("Ingresos Netos Rafa", ["Ingresos netos"], "NORMAL", "CURRENCY"),
        ("Gastos Individuales Rafa", ["Gastos"], "NORMAL", "CURRENCY"),
        ("Cuota Rafa a Familia", [], "NORMAL", "CURRENCY"),
        ("Ingresos Brutos Cris", [], "NORMAL", "CURRENCY"),
        ("Ingresos Netos Cris", [], "NORMAL", "CURRENCY"),
        ("Gastos Individuales Cris", [], "NORMAL", "CURRENCY"),
        ("Cuota Cris a Familia", [], "NORMAL", "CURRENCY"),
        ("Gastos Familia (Común)", ["Gastos comun"], "NORMAL", "CURRENCY"),
        ("", [], "EMPTY", ""),
        
        ("📝 BITÁCORA CONTABLE", [], "HEADER_MAIN", "TEXT"),
        ("Notas del Mes", [], "NORMAL", "TEXT"),
    ]

    # STYLES
    bg_main = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Blue
    bg_sub = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light Blue
    bg_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Column Headers
    bg_total = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Green Total Columns
    
    font_white_bold = Font(bold=True, color="FFFFFF")
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Build Headers (Rows 1, 2, 3)
    ws.cell(row=3, column=1, value="CONCEPTO")
    ws.cell(row=3, column=2, value="Nº CUENTA")
    
    for c in range(1, 3):
        ws.cell(row=3, column=c).fill = bg_header
        ws.cell(row=3, column=c).font = font_white_bold
        ws.cell(row=3, column=c).alignment = align_center

    # Organize columns by Calendar Year
    col_idx = 3 # Start writing months at column C
    year_groups = {} # year -> [col_indices]
    
    col_mapping = {} # old_col -> new_col
    for old_col, y_num, m_num, nice_name in cols_data:
        ws.cell(row=3, column=col_idx, value=nice_name)
        ws.cell(row=3, column=col_idx).fill = bg_header
        ws.cell(row=3, column=col_idx).font = font_white_bold
        ws.cell(row=3, column=col_idx).alignment = align_center
        
        ws.cell(row=2, column=col_idx, value=1.18) # Default divisor just in case
        
        col_mapping[old_col] = col_idx
        if y_num not in year_groups:
            year_groups[y_num] = []
        year_groups[y_num].append(col_idx)
        
        col_idx += 1
        
        # Check if this is December, if so, append a TOTAL column for that year
        # Or if it's the last month of that year available
        # Wait, better to append TOTAL after all months of a year are placed
        
    # Re-arranging so that TOTAL comes after the year
    # Actually, the loop above already laid them out. We need to insert TOTAL columns.
    # Let's rebuild the column layout more cleanly.
    
    ws.delete_rows(1, 4) # Clear and restart header building
    ws.cell(row=1, column=1, value="CONCEPTO").fill = bg_header
    ws.cell(row=1, column=2, value="Nº CUENTA").fill = bg_header
    
    col_idx = 3
    col_mapping = {}
    
    # Sort years
    years = sorted(list(set([y for _, y, _, _ in cols_data])))
    
    for yr in years:
        months_in_yr = [x for x in cols_data if x[1] == yr]
        for old_col, y_num, m_num, nice_name in months_in_yr:
            ws.cell(row=1, column=col_idx, value=nice_name)
            ws.cell(row=1, column=col_idx).fill = bg_header
            ws.cell(row=1, column=col_idx).font = font_white_bold
            
            col_mapping[old_col] = col_idx
            col_idx += 1

    # Freeze Panes
    ws.freeze_panes = 'C2'

    # Now write the rows
    current_row = 2
    for new_name, search_terms, row_type, data_type in structure:
        ws.cell(row=current_row, column=1, value=new_name)
        
        cell_A = ws.cell(row=current_row, column=1)
        if row_type == "HEADER_MAIN":
            cell_A.fill = bg_main
            cell_A.font = font_white_bold
        elif row_type == "HEADER_SUB" or row_type == "FORMULA":
            cell_A.fill = bg_sub
            cell_A.font = font_bold
            
        # Add a blank for Nº Cuenta unless it's a header
        if row_type not in ["HEADER_MAIN", "EMPTY"]:
            ws.cell(row=current_row, column=2).alignment = align_center
            
        # Copy values from old columns to new columns
        if row_type not in ["HEADER_MAIN", "EMPTY"] and search_terms:
            old_r = find_old_row(search_terms, old_row_map)
            if old_r:
                for old_col, new_col in col_mapping.items():
                    val = old_ws.cell(row=old_r, column=old_col).value
                    if val is not None:
                        c = ws.cell(row=current_row, column=new_col, value=val)
                        # Format
                        if data_type == "CURRENCY": c.number_format = '#,##0.00'
                        elif data_type == "PERCENT": c.number_format = '0.00%'

        current_row += 1

    # Format Column Widths
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 15
    for c in range(3, col_idx):
        ws.column_dimensions[get_column_letter(c)].width = 15

    # Save
    wb.save(new_path)
    print(f"Generated V3 Official successfully at: {new_path}")

if __name__ == '__main__':
    build_v3()
